"""
POST /api/extract-supplier
Body: { "file_base64": "<base64>", "filename": "statement.pdf" }
(also accepts the older key "pdf_base64" for backward compatibility)

Deterministic, format-agnostic supplier-statement parser - no LLM, no
external calls. Every supplier sends statements in their own layout and
their own file format, so this endpoint:
  1. sniffs the actual file type from its bytes (PDF / XLSX / CSV /
     plain text) rather than trusting the filename or content-type,
  2. routes to a format-specific reader that turns the file into plain
     rows of cells,
  3. runs the SAME column-keyword header detection and row-building
     logic against those cells regardless of source format, so adding
     a synonym to COLUMN_KEYWORDS improves every format at once.

Row id priority: an INV/C-N number inside the description (this is the
number that also appears on our statement as "Inv#"), else whichever
id-like column the sheet provides (invoice number > check/voucher number
> transaction id). Dates in "Jan 12, 2026", "12/01/2026", or native Excel
date form are all normalized to YYYY-MM-DD.
"""

from http.server import BaseHTTPRequestHandler
import base64
import csv
import io
import json
import re

# pdfplumber extracts right-to-left (Arabic/Hebrew) text runs in reversed
# character order - this restores correct reading order for just those
# runs, leaving any interleaved English/numbers/punctuation untouched.
ARABIC_RUN_RE = re.compile(r"[\u0600-\u06FF\u0750-\u077F][\u0600-\u06FF\u0750-\u077F\s]*[\u0600-\u06FF\u0750-\u077F]|[\u0600-\u06FF\u0750-\u077F]")


def fix_bidi_text(text):
    return ARABIC_RUN_RE.sub(lambda m: m.group(0)[::-1], text)

import datetime as dt

import pdfplumber
import openpyxl

BUILD_TAG = "2026-08-26-remove-pdc-skipword"

# ------------------------------------------------------------ shared vocab

COLUMN_KEYWORDS = {
    "date": ["date", "invc", "تاريخ", "التاريخ"],
    "id": ["ref", "reference", "invoice", "inv", "voucher", "vch", "vnum", "doc",
           "document", "no", "number", "num", "trx", "nb", "id", "check",
           "chk", "cheque", "رقم", "سند", "مستند", "فاتورة", "المرجع", "مرجع"],
    "description": ["description", "details", "narration", "particulars",
                    "memo", "remarks", "name", "بيان", "البيان", "التفاصيل",
                    "الوصف", "ملاحظات"],
    "debit": ["debit", "debitcu", "dr", "مدين"],
    "credit": ["credit", "creditcu", "cr", "دائن"],
    "balance": ["balance", "balcu", "bal", "رصيد", "الرصيد"],
}

# id-like columns ranked by how reliably they match OUR invoice numbers;
# used only for spreadsheet parsing, where several id columns can coexist
# on one row (e.g. "Trans Id", "Invoice number", "Check Num" all at once)
ID_COLUMN_PRIORITY = ["invoice", "voucher", "vch", "vnum", "check", "chk", "cheque",
                       "doc", "document", "trx", "id",
                       "ref", "reference",
                       "no", "number", "num", "nb"]

OPENING_WORDS = ("opening", "b/f", "b.f", "brf", "brought", "balance until",
                 "افتتاحي", "سابق", "مدور", "رصيد اول")
CLOSING_WORDS = ("closing", "c/f", "c.f", "carried", "total", "grand", "movement",
                 "ending balance", "end date", "balance as at",
                 "اجمالي", "إجمالي", "المجموع", "ختامي", "نهائي")
SKIP_WORDS = ("statement", "page ", "page:", "printed", "tel:", "fax:",
              "p.o.box", "www.", "@")
# "pdc" (postdated cheques) used to be in this list to skip a trailing
# PDC table after the closing balance - removed because it silently
# skipped every payment row on a real statement whose own payment
# reference numbers are prefixed "JPDC-..." ("pdc" matched as a
# substring inside "jpdc"), dropping every JPMT row in the entire
# document. The "stop after closing balance" guard in
# parse_words_strategy/parse_tables_strategy already excludes the
# trailing PDC section by a precise mechanism (nothing after the closing
# balance line is ever parsed), making this keyword both redundant and
# actively dangerous.

# A page-footer print timestamp ("Date 07/08/2026 Time 2:41:43PM") looks
# just enough like a data row (has a date, "Time" isn't in any keyword
# list) to slip through as a phantom $0/$0 transaction. Harmless to
# matching (compare.py already drops all-zero rows on both sides) but
# pure noise - skip it generally by its distinctive clock-time shape
# rather than any one supplier's exact footer wording.
PRINT_TIMESTAMP_RE = re.compile(r"\b\d{1,2}:\d{2}:\d{2}\s*[AaPp][Mm]\b")

AMOUNT_RE = re.compile(r"^\(?-?(?:[\d,]+(?:\.\d+)?|\.\d+)\)?(CR|DR)?$", re.IGNORECASE)
NUM_DATE_RE = re.compile(r"(\d{1,4})\s*[/\-.]\s*(\d{1,2})\s*[/\-.]\s*(\d{1,4})")
MONTH_DATE_RE = re.compile(
    r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?\s+(\d{1,2})\s*,?\s+(\d{4})",
    re.IGNORECASE)
MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
          "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
REF_RE = re.compile(r"(?:INV|C\s*R?\s*/?\s*N)\s*[:#]?\s*([\w/-]+)", re.IGNORECASE)


def is_valid_col_map(mapped):
    """A header mapping only counts if debit/credit/balance actually sit in
    DIFFERENT physical columns, and separately from date/id/description.
    Some PDFs draw horizontal row rules with no vertical column dividers at
    all, so a table-detection pass can find "rows" but only ever returns
    one wide cell per row; naively matching keywords against that single
    cell's text can make multiple fields appear to share one column index.
    That's not a real column split, so reject it and let the caller fall
    back to a strategy that reads actual word positions instead."""
    if len(mapped) < 3 or not any(c in mapped for c in ("debit", "credit", "balance")):
        return False
    return len(set(mapped.values())) >= 3


def undouble_chars(word):
    """Some PDF exports fake a bold header by drawing each character
    twice at a near-identical position, which pdfplumber's text
    extraction reads as every character literally repeated ('Number'
    becomes 'NNuummbbeerr'). Detect that specific pattern within a single
    whitespace-separated word and collapse it back to the original -
    ordinary text (including genuine double letters like 'Statement')
    won't match this check and passes through unchanged."""
    if len(word) < 2 or len(word) % 2 != 0:
        return word
    if all(word[i] == word[i + 1] for i in range(0, len(word), 2)):
        return word[0::2]
    return word


def match_column(text):
    t = str(text or "").strip().lower().strip(":.")
    for col, words in COLUMN_KEYWORDS.items():
        if t in words:
            return col
    parts = [p.strip(":.") for p in t.split()]
    for col, words in COLUMN_KEYWORDS.items():
        if any(p in words for p in parts):
            return col
    # retry against a de-doubled version, in case this cell hit the
    # fake-bold rendering artifact described above
    fixed_parts = [undouble_chars(p) for p in parts]
    if fixed_parts != parts:
        for col, words in COLUMN_KEYWORDS.items():
            if any(p in words for p in fixed_parts):
                return col
    return None


def id_column_rank(header_text):
    """How specific an id-like header is, for picking the best of several
    id columns on one spreadsheet row. Lower is more trustworthy."""
    t = str(header_text or "").strip().lower()
    for i, kw in enumerate(ID_COLUMN_PRIORITY):
        if kw in t:
            return i
    # retry against a de-doubled version (see undouble_chars / match_column)
    t_fixed = " ".join(undouble_chars(p) for p in t.split())
    if t_fixed != t:
        for i, kw in enumerate(ID_COLUMN_PRIORITY):
            if kw in t_fixed:
                return i
    return len(ID_COLUMN_PRIORITY)


def parse_amount(token):
    token = str(token).strip()
    # strip common currency symbols/codes that can appear anywhere in the
    # token, including inside parentheses for negatives ("($1,085.77)")
    cleaned = re.sub(r"(?i)us\$|usd|\$", "", token).strip()
    m = AMOUNT_RE.match(cleaned)
    if not m:
        return None
    s = cleaned[: len(cleaned) - 2] if m.group(1) else cleaned
    s = s.replace(",", "").replace("(", "-").replace(")", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_amount_cell(raw_cell):
    """A debit/credit/balance cell can end up with more than one amount-
    like token in it - e.g. a description that restates an amount inline
    ("...4619/102/02- 24.40 $") can sit right at a column boundary, so
    both that memo number and the real column value land in the same
    cell ("24.40 $ 24.40"), which fails to parse as a single amount and
    silently becomes 0. assign_columns joins a cell's words in
    left-to-right x-order, so when the whole cell doesn't parse cleanly,
    the LAST token is the one physically closest to the real column
    position and by far the more likely genuine value."""
    if raw_cell == "":
        return None
    val = parse_amount(raw_cell)
    if val is not None:
        return val
    parts = raw_cell.split()
    if len(parts) > 1:
        return parse_amount(parts[-1])
    return None


# Per-document date convention ('DMY' or 'MDY'), detected once per parsed
# file from unambiguous dates it contains (see detect_date_convention).
# Only used for the numeric-slash date form when both parts could be either
# a day or a month (e.g. "4/1/2026") - a date like "28/05/2026" or
# "5/15/2026" is unambiguous regardless of this setting, since one of the
# two numbers can't possibly be a month.
_DATE_CONVENTION = "DMY"


def detect_date_convention(text):
    """Scan a document's full text for numeric dates and infer whether it
    uses day-first or month-first ordering, based on any date where one of
    the first two numbers exceeds 12 (which can only be a day, never a
    month). Defaults to day-first (our own convention) when the document
    has no unambiguous evidence either way."""
    dmy_evidence = mdy_evidence = 0
    for m in NUM_DATE_RE.finditer(text):
        a, b = int(m.group(1)), int(m.group(2))
        if a > 31 or b > 31:
            continue  # not a plausible day/month pair at all (likely a year-first date)
        if a > 12 and b <= 12:
            dmy_evidence += 1
        elif b > 12 and a <= 12:
            mdy_evidence += 1
    return "MDY" if mdy_evidence > dmy_evidence else "DMY"


def extract_and_strip_date(text):
    """Find a date anywhere in text, returning (iso_date_or_None,
    remaining_text_with_the_date_removed). Used both by find_date() and by
    the "no separate date column" fallback in build_row(), where a date
    that shares a column with other fields (id, type code, etc.) needs to
    be pulled out before the rest of that cell's content can be parsed."""
    text = str(text or "")
    m = MONTH_DATE_RE.search(text)
    if m:
        mo = MONTHS[m.group(1).lower()[:3]]
        d, y = int(m.group(2)), int(m.group(3))
        if 1 <= d <= 31:
            remaining = (text[:m.start()] + text[m.end():]).strip()
            return f"{y:04d}-{mo:02d}-{d:02d}", remaining
    m = NUM_DATE_RE.search(text)
    if m:
        a, b, c = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if a > 31:
            y, mo, d = a, b, c
        elif a > 12 and b <= 12:
            d, mo, y = a, b, c            # unambiguous: a can't be a month
        elif b > 12 and a <= 12:
            mo, d, y = a, b, c            # unambiguous: b can't be a month
        elif _DATE_CONVENTION == "MDY":
            mo, d, y = a, b, c
        else:
            d, mo, y = a, b, c
        if y < 100:
            y += 2000
        if 1 <= d <= 31 and 1 <= mo <= 12 and 1900 < y < 2100:
            remaining = (text[:m.start()] + text[m.end():]).strip()
            return f"{y:04d}-{mo:02d}-{d:02d}", remaining
    return None, text


def find_date(text):
    """Search a cell for a date in 'Jan 12, 2026' or numeric form -> ISO."""
    date, _ = extract_and_strip_date(text)
    return date


def extract_row_id(id_cell, description):
    m = REF_RE.search(str(description or ""))
    if m:
        # "INV 26-2021" style refs pack a short prefix (year, branch code)
        # and the real invoice number together; take the longest digit run
        # rather than just whatever comes immediately after INV/C-N, so a
        # short prefix doesn't win over the actual number that follows it.
        runs = re.findall(r"\d+", m.group(1))
        if runs:
            return max(runs, key=len)
    for tok in str(id_cell or "").split():
        if re.search(r"\d", tok) and not find_date(tok):
            return tok
    return ""


def build_row(cells, raw_low):
    date = find_date(cells.get("date", "")) or ""
    # Some suppliers don't print a separate "Date" column header at all -
    # the date is just an unlabeled leading field that ends up sharing
    # whatever column comes first (often "id"). When there's genuinely no
    # dedicated date column, search for an embedded date there instead,
    # and strip it out so it doesn't pollute id extraction afterward.
    if not date and not cells.get("date") and cells.get("id"):
        found_date, remainder = extract_and_strip_date(cells["id"])
        if found_date:
            date = found_date
            cells = dict(cells)
            cells["id"] = remainder
    debit = parse_amount_cell(cells.get("debit", ""))
    credit = parse_amount_cell(cells.get("credit", ""))
    balance = parse_amount_cell(cells.get("balance", ""))

    if any(k in raw_low for k in OPENING_WORDS):
        row_type = "opening_balance"
    elif any(k in raw_low for k in CLOSING_WORDS):
        row_type = "closing_balance"
    else:
        row_type = "transaction"

    if not (date or any(v is not None for v in (debit, credit, balance))):
        return None
    # Note: a dateless "transaction" row is NOT dropped here - the caller
    # gets first chance to fill in a date inherited from the previous row
    # (some suppliers print a date once per batch of line items and leave
    # it blank on the rest of that batch), and only drops the row if it's
    # still dateless after that - see parse_tables_strategy /
    # parse_words_strategy.
    desc = fix_bidi_text(cells.get("description", ""))
    return {
        "date": date,
        "id": extract_row_id(cells.get("id", ""), desc) if row_type == "transaction" else "",
        "description": desc,
        # sign is presentation-only (some suppliers print credits negative)
        "debit": abs(debit) if debit is not None else 0.0,
        "credit": abs(credit) if credit is not None else 0.0,
        "balance": balance if balance is not None else 0.0,
        "row_type": row_type,
    }


def is_continuation(cells):
    if find_date(cells.get("date", "")):
        return False
    if any(parse_amount(cells.get(c, "")) is not None for c in ("debit", "credit", "balance") if cells.get(c, "") != ""):
        return False
    return bool(cells.get("description") or cells.get("id"))


# ================================================================== PDF

def group_lines(page):
    words = page.extract_words(x_tolerance=1.5, y_tolerance=2.0, keep_blank_chars=False)
    words.sort(key=lambda w: (w["top"], w["x0"]))
    lines, cur, cur_top = [], [], None
    for w in words:
        if cur_top is None or abs(w["top"] - cur_top) <= 2.5:
            cur.append(w)
            if cur_top is None:
                cur_top = w["top"]
        else:
            lines.append(sorted(cur, key=lambda x: x["x0"]))
            cur, cur_top = [w], w["top"]
    if cur:
        lines.append(sorted(cur, key=lambda x: x["x0"]))
    return lines


def find_header(lines):
    best, best_count = None, 0
    for line in lines:
        cols = {}
        for w in line:
            col = match_column(w["text"])
            if col and col not in cols:
                cols[col] = (w["x0"] + w["x1"]) / 2.0
        if len(cols) >= 3 and any(c in cols for c in ("debit", "credit", "balance")):
            if len(cols) > best_count:
                best, best_count = cols, len(cols)
    return best


def build_intervals(anchors, page_width):
    ordered = sorted(anchors.items(), key=lambda kv: kv[1])
    intervals = []
    for idx, (col, x) in enumerate(ordered):
        left = 0 if idx == 0 else (ordered[idx - 1][1] + x) / 2.0
        right = page_width if idx == len(ordered) - 1 else (x + ordered[idx + 1][1]) / 2.0
        intervals.append((col, left, right))
    return intervals


def assign_columns(line, intervals):
    cells = {col: [] for col, _, _ in intervals}
    for w in line:
        center = (w["x0"] + w["x1"]) / 2.0
        for col, left, right in intervals:
            if left <= center < right:
                cells[col].append(w["text"].strip())
                break
    return {col: " ".join(v).strip() for col, v in cells.items()}


def word_column(w, intervals):
    """Which column a word's x-position actually falls into, per the same
    intervals assign_columns() uses."""
    center = (w["x0"] + w["x1"]) / 2.0
    for col, left, right in intervals:
        if left <= center < right:
            return col
    return None


def parse_words_strategy(pdf):
    rows, warnings = [], []
    anchors = None
    # Once a closing-balance line is seen, NOTHING after it is ever a real
    # transaction - it's always trailing statement metadata (postdated
    # check lists, promissory notes, signature blocks, disclaimers). Seen
    # for real: a "PDC" (postdated cheques) table after the closing
    # balance got misparsed as a transaction row, with the check's Due
    # Date column mistaken for an id and its Amount column mistaken for a
    # debit - a completely different table structure that happens to also
    # have dates and amounts in columns.
    seen_closing = False
    for page_no, page in enumerate(pdf.pages, start=1):
        if seen_closing:
            break
        lines = group_lines(page)
        page_anchors = find_header(lines)
        if page_anchors:
            anchors = page_anchors
        if not anchors:
            continue
        intervals = build_intervals(anchors, page.width)

        prev_was_data = False
        prev_date = ""
        for line in lines:
            if seen_closing:
                break
            raw = " ".join(w["text"] for w in line).strip()
            if not raw:
                continue
            raw_low = raw.lower()
            # A repeated header line has its keyword words sitting in the
            # SAME columns those keywords represent (e.g. "Debit" printed
            # at the debit column's x-position). Matching by keyword alone
            # is not enough - a data row's Type code or description can
            # legitimately contain header vocabulary too (e.g. a supplier
            # whose transaction type is literally "INV", or whose
            # description reads "INV No. 12345" - "inv"/"no" both match
            # the id-column keyword list even though they're sitting in
            # the type/description columns, not the id column). Only count
            # a hit when the word's actual x-position column agrees with
            # what the keyword implies.
            header_hits = sum(
                1 for w in line
                if match_column(w["text"]) and match_column(w["text"]) == word_column(w, intervals)
            )
            if header_hits >= 3:
                prev_was_data = False
                continue
            if any(s in raw_low for s in SKIP_WORDS) or PRINT_TIMESTAMP_RE.search(raw):
                prev_was_data = False
                continue

            cells = assign_columns(line, intervals)
            row = build_row(cells, raw_low)
            if row:
                # See parse_tables_strategy: some suppliers print a date
                # once per batch of line items and leave it blank on the
                # rest of that batch.
                if not row["date"] and prev_date and row["row_type"] == "transaction":
                    row["date"] = prev_date
                if row["row_type"] == "transaction" and not row["date"]:
                    prev_was_data = False
                    continue  # still no usable date - a genuine stray line
                if row["date"]:
                    prev_date = row["date"]
                rows.append(row)
                prev_was_data = True
                if row["row_type"] == "closing_balance":
                    seen_closing = True
            elif prev_was_data and is_continuation(cells) and rows:
                extra = (cells.get("description", "") + " " + cells.get("id", "")).strip()
                rows[-1]["description"] = (rows[-1]["description"] + " " + extra).strip()
                if rows[-1]["row_type"] == "transaction" and not rows[-1]["id"]:
                    rows[-1]["id"] = extract_row_id("", rows[-1]["description"])
            elif len(raw) > 3:
                warnings.append(f"page {page_no}: unclassified line: {raw[:120]}")
                prev_was_data = False
    return rows, warnings, anchors is not None


def find_external_header_mapping(page, table):
    """When a table has no header row of its own (some invoicing systems
    print the column labels as free text directly above the ruled grid,
    outside the table pdfplumber detects), look for that header line
    among the words sitting above the table's top edge, and map each
    matched keyword to whichever of the table's own column x-ranges it
    falls inside. Returns a {field: column_index} dict, or None."""
    if not table.rows:
        return None
    first_row_cells = table.rows[0].cells
    if not first_row_cells or any(c is None for c in first_row_cells):
        return None
    col_ranges = [(c[0], c[2]) for c in first_row_cells]  # (x0, x1) per column

    top_of_table = table.bbox[1]
    words = page.extract_words(x_tolerance=1.5, y_tolerance=2.0, keep_blank_chars=False)
    candidates = [w for w in words if top_of_table - 80 <= w["top"] < top_of_table]
    candidates.sort(key=lambda w: w["top"])

    lines, cur, cur_top = [], [], None
    for w in candidates:
        if cur_top is None or abs(w["top"] - cur_top) <= 2.5:
            cur.append(w)
            cur_top = w["top"] if cur_top is None else cur_top
        else:
            lines.append(cur)
            cur, cur_top = [w], w["top"]
    if cur:
        lines.append(cur)

    best_map, best_count = None, 0
    for line in lines:
        mapping = {}
        for w in line:
            col = match_column(w["text"])
            if not col:
                continue
            wx = (w["x0"] + w["x1"]) / 2
            for idx, (x0, x1) in enumerate(col_ranges):
                if x0 <= wx < x1:
                    if col not in mapping:
                        mapping[col] = idx
                    break
        if len(mapping) > best_count and is_valid_col_map(mapping):
            best_map, best_count = mapping, len(mapping)
    return best_map


def parse_tables_strategy(pdf):
    rows, warnings = [], []
    found_any = False
    last_col_map, last_id_candidates, last_col_count = None, [], None
    seen_closing = False  # see the same guard/comment in parse_words_strategy
    prev_date = ""  # persists across separate table regions AND pages, since
                    # a single logical table can get split into multiple
                    # detected regions (e.g. around a page's mid-content
                    # boilerplate), and a date-fill-down row right at that
                    # boundary shouldn't lose its inherited date just
                    # because pdfplumber happened to see it as a new table.
    for page in pdf.pages:
        if seen_closing:
            break
        for table in page.find_tables():
            if seen_closing:
                break
            table_rows = table.extract()
            if not table_rows:
                continue
            this_col_count = len(table_rows[0]) if table_rows[0] else 0

            col_map, id_candidates, data_rows = None, [], table_rows
            # first choice: a header row printed inside the table itself.
            # map_sheet_header ranks every id-like column (a table can have
            # several - "Accounting Ref.", "Document Nbr." - and the one
            # that's actually populated and matches our own numbering
            # varies per supplier) instead of only keeping whichever
            # appeared first.
            for i, cells in enumerate(table_rows[:10]):
                cells_norm = [(c or "").replace("\n", " ").strip() for c in cells]
                mapped, ids = map_sheet_header(cells_norm)
                if is_valid_col_map(mapped):
                    col_map, id_candidates, data_rows = mapped, ids, table_rows[i + 1:]
                    break
            # second choice: header printed as free text above the ruled grid
            if col_map is None:
                col_map = find_external_header_mapping(page, table)
                data_rows = table_rows  # every extracted row is data in this case
            # third choice: a later page of the same multi-page statement,
            # where the header (in either form above) was only printed once
            # on page 1 and doesn't repeat - reuse the last mapping as long
            # as this table has the same number of columns, since a
            # differently-shaped table is almost certainly something else
            # (e.g. an address block) and applying a stale map to it would
            # silently produce garbage rows.
            if col_map is None and last_col_map is not None and this_col_count == last_col_count:
                col_map, id_candidates = last_col_map, last_id_candidates
                data_rows = table_rows
            if col_map is None:
                continue

            last_col_map, last_id_candidates, last_col_count = col_map, id_candidates, this_col_count
            found_any = True
            prev_was_data = False
            for cells in data_rows:
                if seen_closing:
                    break
                cells = [(c or "").replace("\n", " ").strip() for c in cells]

                def get(field):
                    i = col_map.get(field, -1)
                    return cells[i] if 0 <= i < len(cells) else ""

                celld = {f: get(f) for f in COLUMN_KEYWORDS}
                # pick the first ranked id-candidate column that's actually
                # populated on THIS row - some suppliers group several line
                # items under one shared reference, printed only once, so a
                # later, more specific column (e.g. a per-line document
                # number) may be the one that's really filled in here.
                for _, idx in id_candidates:
                    if idx < len(cells) and cells[idx]:
                        celld["id"] = cells[idx]
                        break
                raw_low = " ".join(cells).lower()
                raw_joined = " ".join(cells)
                if any(s in raw_low for s in SKIP_WORDS) or PRINT_TIMESTAMP_RE.search(raw_joined):
                    prev_was_data = False
                    continue
                row = build_row(celld, raw_low)
                if row and not row["date"] and row["row_type"] == "transaction":
                    # The winning id column is correct but has no date in
                    # it - some layouts put date+type combined in a
                    # DIFFERENT, lower-priority id-like column (e.g. a
                    # "Number" column that's actually date+type, while the
                    # real invoice number lives in "Vnum"). Scan every
                    # other id-candidate cell for an embedded date before
                    # giving up.
                    for _, idx in id_candidates:
                        if idx < len(cells) and cells[idx]:
                            found_date, _ = extract_and_strip_date(cells[idx])
                            if found_date:
                                row["date"] = found_date
                                break
                if row:
                    # Some suppliers print the date once per batch of line
                    # items and leave it blank on subsequent lines in that
                    # batch - inherit the last seen date rather than
                    # discarding a real transaction just because its own
                    # date cell was blank.
                    if not row["date"] and prev_date and row["row_type"] == "transaction":
                        row["date"] = prev_date
                    if row["row_type"] == "transaction" and not row["date"]:
                        prev_was_data = False
                        continue  # still no usable date - a genuine stray line
                    if row["date"]:
                        prev_date = row["date"]
                    rows.append(row)
                    prev_was_data = True
                    if row["row_type"] == "closing_balance":
                        seen_closing = True
                elif prev_was_data and is_continuation(celld) and rows:
                    extra = (celld.get("description", "") + " " + celld.get("id", "")).strip()
                    if extra:
                        rows[-1]["description"] = (rows[-1]["description"] + " " + extra).strip()
    return rows, warnings, found_any


def parse_pdf(file_bytes):
    global _DATE_CONVENTION
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        pages = len(pdf.pages)
        total_chars = sum(len(p.chars) for p in pdf.pages)
        if total_chars < 20:
            raise ValueError(
                "This PDF has no text layer (it is a scan/image). "
                "It needs OCR before it can be parsed without an LLM."
            )
        full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        _DATE_CONVENTION = detect_date_convention(full_text)

        t_rows, t_warn, t_found = parse_tables_strategy(pdf)
        w_rows, w_warn, w_found = parse_words_strategy(pdf)

    # Ruled-table extraction reads each printed cell directly and is immune
    # to the column-bleed that can happen with the word-position strategy
    # (e.g. an address like "CHTAURA-" running into the next column with no
    # gap). Prefer it whenever the PDF actually draws table lines - but
    # "t_found" only means pdfplumber's detector matched SOMETHING; its
    # whitespace/alignment heuristics can false-positive on a PDF with no
    # real ruled grid at all, swallowing every row into one giant merged
    # cell (seen for real: 22 transactions collapsed into a single
    # "opening_balance" row because a genuine ruled table doesn't exist on
    # that PDF). A real ruled-table extraction should never produce fewer
    # actual transaction rows than the word-position fallback finds on the
    # same page, so use that as the sanity check rather than trusting
    # "t_found" alone.
    t_txn_count = sum(1 for r in t_rows if r.get("row_type") == "transaction")
    w_txn_count = sum(1 for r in w_rows if r.get("row_type") == "transaction")
    if t_found and t_rows and t_txn_count >= w_txn_count:
        strategy, rows, warnings = "table", t_rows, t_warn
    elif w_found:
        strategy, rows, warnings = "words", w_rows, w_warn
    elif t_found and t_rows:
        strategy, rows, warnings = "table", t_rows, t_warn
    else:
        raise ValueError(
            "Could not find a recognizable column header (Date/Debit/Credit/"
            "Balance or Arabic equivalents) in this PDF. This supplier's "
            "format needs a keyword added to COLUMN_KEYWORDS."
        )
    warnings = [f"[pdf/{strategy}] {w}" for w in warnings]
    if not rows:
        warnings.append(f"[pdf/{strategy}] Header found but no data rows extracted.")
    return rows, warnings, {"pages": pages}


# ================================================================== XLSX

def cell_to_text(v):
    """Excel cells already carry typed values (dates, numbers) - normalize
    everything to the string form the shared header/amount matchers expect,
    except dates which are converted straight to ISO to avoid ambiguity."""
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def map_sheet_header(header_row):
    """Map each column index to a field name. Unlike the PDF path, a
    spreadsheet can have several id-like columns on one row (Trans Id,
    Invoice number, Check Num); keep all of them ranked so build_sheet_row
    can pick the most specific one actually populated on each row, instead
    of only keeping whichever appeared first."""
    mapped = {}       # field -> single column index (date/description/debit/credit/balance)
    id_candidates = []  # list of (rank, column index) for every id-like header
    for idx, cell in enumerate(header_row):
        col = match_column(cell)
        if col == "id":
            id_candidates.append((id_column_rank(cell), idx))
        elif col and col not in mapped:
            mapped[col] = idx
    id_candidates.sort()
    return mapped, id_candidates


def build_sheet_row(row_values, mapped, id_candidates):
    def get(field):
        i = mapped.get(field, -1)
        return cell_to_text(row_values[i]) if 0 <= i < len(row_values) else ""

    cells = {f: get(f) for f in ("date", "description", "debit", "credit", "balance")}
    # pick the first (highest-priority) id column that actually has a value
    # on this specific row, since e.g. "Check Num" is usually blank except
    # on cheque rows while "Invoice number" is blank on those same rows
    id_text = ""
    for _, idx in id_candidates:
        if idx < len(row_values):
            v = cell_to_text(row_values[idx])
            if v:
                id_text = v
                break
    cells["id"] = id_text

    raw_low = " ".join(cell_to_text(v) for v in row_values if v is not None).lower()
    return build_row(cells, raw_low), cells


def parse_xlsx(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    rows, warnings = [], []
    sheets_used = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        header_idx, mapped, id_candidates = None, None, None
        for i, r in enumerate(all_rows[:5]):
            m, ids = map_sheet_header(r)
            if len(m) + (1 if ids else 0) >= 3 and any(c in m for c in ("debit", "credit", "balance")):
                header_idx, mapped, id_candidates = i, m, ids
                break
        if header_idx is None:
            continue  # this sheet doesn't look like a ledger - skip quietly

        sheets_used += 1
        prev_row_ref = None
        for r in all_rows[header_idx + 1:]:
            if r is None or all(v is None for v in r):
                continue
            row, cells = build_sheet_row(r, mapped, id_candidates)
            if row:
                rows.append(row)
                prev_row_ref = rows[-1]
            elif prev_row_ref is not None and is_continuation(cells):
                extra = (cells.get("description", "") + " " + cells.get("id", "")).strip()
                if extra:
                    prev_row_ref["description"] = (prev_row_ref["description"] + " " + extra).strip()

    if sheets_used == 0:
        raise ValueError(
            "Could not find a recognizable column header (Date/Debit/Credit/"
            "Balance or similar) in any sheet of this workbook. This "
            "supplier's format needs a keyword added to COLUMN_KEYWORDS."
        )
    if not rows:
        warnings.append("[xlsx] Header(s) found but no data rows extracted.")
    return rows, warnings, {"sheets": len(wb.sheetnames), "sheets_used": sheets_used}


# ================================================================== CSV

def parse_csv(file_bytes):
    global _DATE_CONVENTION
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1256", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Could not decode this CSV file as text.")
    _DATE_CONVENTION = detect_date_convention(text)

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = list(csv.reader(io.StringIO(text), dialect))
    reader = [r for r in reader if any(c.strip() for c in r)]
    if not reader:
        raise ValueError("This CSV file has no rows.")

    header_idx, mapped, id_candidates = None, None, None
    for i, r in enumerate(reader[:5]):
        m, ids = map_sheet_header(r)
        if len(m) + (1 if ids else 0) >= 3 and any(c in m for c in ("debit", "credit", "balance")):
            header_idx, mapped, id_candidates = i, m, ids
            break
    if header_idx is None:
        raise ValueError(
            "Could not find a recognizable column header (Date/Debit/Credit/"
            "Balance or similar) in this CSV. This supplier's format needs "
            "a keyword added to COLUMN_KEYWORDS."
        )

    rows, warnings = [], []
    prev_row_ref = None
    for r in reader[header_idx + 1:]:
        row, cells = build_sheet_row(r, mapped, id_candidates)
        if row:
            rows.append(row)
            prev_row_ref = rows[-1]
        elif prev_row_ref is not None and is_continuation(cells):
            extra = (cells.get("description", "") + " " + cells.get("id", "")).strip()
            if extra:
                prev_row_ref["description"] = (prev_row_ref["description"] + " " + extra).strip()

    if not rows:
        warnings.append("[csv] Header found but no data rows extracted.")
    return rows, warnings, {}


# ================================================================== dispatch

def sniff_format(file_bytes, filename):
    """Identify the real file format from its bytes first (magic numbers
    can't be spoofed by a wrong extension), falling back to the filename
    extension only when the bytes are inconclusive (plain text formats)."""
    if file_bytes.startswith(b"%PDF"):
        return "pdf"
    if file_bytes.startswith(b"PK\x03\x04"):
        return "xlsx"     # xlsx/xlsm are zip containers
    if file_bytes.startswith(b"\xd0\xcf\x11\xe0"):
        raise ValueError(
            "This looks like a legacy .xls file (pre-2007 Excel format). "
            "Please re-save it as .xlsx or .csv and upload again."
        )
    ext = (filename or "").rsplit(".", 1)[-1].lower() if filename else ""
    if ext in ("csv", "tsv", "txt"):
        return "csv"
    if ext in ("xlsx", "xlsm"):
        return "xlsx"
    if ext == "pdf":
        return "pdf"
    # last resort: does it decode as text at all?
    try:
        file_bytes[:2048].decode("utf-8")
        return "csv"
    except UnicodeDecodeError:
        raise ValueError(
            "Could not identify this file's format. Supported formats: "
            "PDF, Excel (.xlsx), and CSV."
        )


VAT_SUFFIX_RE = re.compile(r"(-v|vat)\s*$", re.IGNORECASE)


def _longest_digit_run(s):
    runs = re.findall(r"\d+", str(s or ""))
    return max(runs, key=len).lstrip("0") if runs else ""


def merge_split_vat_lines(rows):
    """Some suppliers print a transaction's tax as a separate line instead
    of combining it into one total the way we do - either right after the
    base line, or (as seen in practice) in a completely separate block
    later in the same document. Fold any such VAT-marked line into the
    base row sharing its date and normalized id, wherever in the document
    each one appears, so the combined total can match our single figure.
    """
    # Pass 1: pull out every VAT-marked row (order-independent) and
    # accumulate its amount per (date, normalized id) key.
    base_rows, vat_totals = [], {}
    for r in rows:
        if r["row_type"] != "transaction":
            base_rows.append(r)
            continue
        nid = _longest_digit_run(r["id"])
        is_vat = nid and (VAT_SUFFIX_RE.search(r["id"].strip())
                          or VAT_SUFFIX_RE.search(r["description"].strip()))
        if is_vat:
            key = (r["date"], nid)
            e = vat_totals.setdefault(key, {"debit": 0.0, "credit": 0.0, "id": ""})
            e["debit"] += r["debit"]
            e["credit"] += r["credit"]
            if len(r["id"]) > len(e["id"]):
                e["id"] = r["id"]
        else:
            base_rows.append(r)

    applied = set()
    with_vat_applied = []
    for r in base_rows:
        if r.get("row_type") == "transaction":
            key = (r["date"], _longest_digit_run(r["id"]))
            if key[1] and key in vat_totals and key not in applied:
                e = vat_totals[key]
                r = dict(r)
                r["debit"] += e["debit"]
                r["credit"] += e["credit"]
                if len(e["id"]) > len(r["id"]):
                    r["id"] = e["id"]
                applied.add(key)
        with_vat_applied.append(r)
    # a VAT line whose base row was never found (id mismatch, or the base
    # row landed outside this file's date filter) - keep it rather than
    # silently dropping the amount.
    for key, e in vat_totals.items():
        if key not in applied:
            with_vat_applied.append({
                "date": key[0], "id": e["id"], "description": "(VAT line, no matching base row found)",
                "debit": e["debit"], "credit": e["credit"], "balance": 0.0, "row_type": "transaction",
            })

    # Pass 2: some suppliers instead print two ADJACENT lines that both
    # carry the identical full description (no distinct VAT marker) -
    # merge those too, same guard against unrelated same-id batches with
    # differing descriptions.
    merged = []
    for r in with_vat_applied:
        prev = merged[-1] if merged else None
        same_group = (
            prev is not None
            and r.get("row_type") == "transaction" == prev.get("row_type")
            and r["date"] == prev["date"]
            and r["description"].strip() != ""
            and r["description"].strip() == prev["description"].strip()
            and _longest_digit_run(r["id"]) == _longest_digit_run(prev["id"])
            and _longest_digit_run(r["id"]) != ""
        )
        if same_group:
            prev["debit"] += r["debit"]
            prev["credit"] += r["credit"]
            if len(r["id"]) > len(prev["id"]):
                prev["id"] = r["id"]
        else:
            merged.append(dict(r))
    return merged


def parse_supplier_file(file_bytes, filename=None):
    fmt = sniff_format(file_bytes, filename)
    if fmt == "pdf":
        rows, warnings, meta = parse_pdf(file_bytes)
    elif fmt == "xlsx":
        rows, warnings, meta = parse_xlsx(file_bytes)
    else:
        rows, warnings, meta = parse_csv(file_bytes)
    rows = merge_split_vat_lines(rows)
    result = {"rows": rows, "warnings": warnings, "build_tag": BUILD_TAG, "format": fmt}
    result.update(meta)
    return result


class handler(BaseHTTPRequestHandler):
    def _send(self, status, payload):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            data = json.loads(self.rfile.read(length) or b"{}")
            b64 = data.get("file_base64") or data.get("pdf_base64", "")
            filename = data.get("filename", "")
            if "," in b64[:80]:
                b64 = b64.split(",", 1)[1]
            if not b64:
                return self._send(400, {"error": "file_base64 is required."})
            file_bytes = base64.b64decode(b64)
            self._send(200, parse_supplier_file(file_bytes, filename))
        except json.JSONDecodeError:
            self._send(400, {"error": "Invalid JSON body."})
        except ValueError as e:
            self._send(422, {"error": str(e)})
        except Exception as e:  # noqa: BLE001
            self._send(500, {"error": f"Extraction failed: {e}"})
